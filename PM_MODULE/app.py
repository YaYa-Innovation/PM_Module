from flask import Flask, render_template, request, jsonify
import os
from openpyxl import load_workbook

app = Flask(__name__)

def calculate_totals(ws, columns):
    # Implement your logic to calculate column_totals and row_totals
    return {}, {}

@app.route('/')
def list_excel_files():
    current_directory = os.path.dirname(os.path.realpath(__file__))
    files = [f for f in os.listdir(current_directory) if f.endswith('.xlsx')]
    return render_template('index.html', files=files)

@app.route('/sheets/<filename>')
def list_sheets(filename):
    current_directory = os.path.dirname(os.path.realpath(__file__))
    file_path = os.path.join(current_directory, filename)
    wb = load_workbook(file_path, read_only=True)
    sheets = wb.sheetnames
    return render_template('sheets.html', filename=filename, sheets=sheets)

@app.route('/columns/<filename>/<sheet_name>', methods=['GET', 'POST'])
def list_columns(filename, sheet_name):
    current_directory = os.path.dirname(os.path.realpath(__file__))
    file_path = os.path.join(current_directory, filename)
    wb = load_workbook(file_path, read_only=False)
    ws = wb[sheet_name]
    columns = [cell.value for cell in ws[1]]

    if request.method == 'POST':
        data_to_store = [request.form[column] for column in columns]
        row_number = ws.max_row + 1

        for col_num, value in enumerate(data_to_store, start=1):
            ws.cell(row=row_number, column=col_num, value=value)

        wb.save(file_path)

    column_totals, row_totals = calculate_totals(ws, columns)
    all_data = list(ws.iter_rows(values_only=True))
    all_data = [list(row) for row in all_data]

    return render_template('columns.html', filename=filename, sheet_name=sheet_name, columns=columns, column_totals=column_totals, row_totals=row_totals, all_data=all_data)

@app.route('/delete_row/<filename>/<sheet_name>/<int:row_number>', methods=['DELETE'])
def delete_row(filename, sheet_name, row_number):
    current_directory = os.path.dirname(os.path.realpath(__file__))
    file_path = os.path.join(current_directory, filename)

    wb = load_workbook(file_path)
    ws = wb[sheet_name]

    # Check if the row exists before deleting
    if 1 <= row_number <= ws.max_row:
        ws.delete_rows(row_number)
        wb.save(file_path)
        return jsonify({'message': f'Row {row_number} deleted successfully'})
    else:
        return jsonify({'error': f'Row {row_number} does not exist'})

if __name__ == '__main__':
    app.run(debug=True)

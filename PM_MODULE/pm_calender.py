import openpyxl
import os

# Set the directory path where Excel files will be saved
excel_directory = 'D:\\PM_MODULE'

def is_sheet_name_exists(workbook, sheet_name):
    """
    Check if the given sheet name already exists in the workbook.
    """
    return sheet_name in workbook.sheetnames

def create_sheet_in_existing_file(file, sheet_name):
    """
    Create a new sheet in an existing Excel file.
    """
    workbook = openpyxl.load_workbook(file)
    
    while True:
        new_sheet_name = input("Enter sheet name: ")

        if is_sheet_name_exists(workbook, new_sheet_name):
            print(f"The sheet name '{new_sheet_name}' already exists. Please enter a different name.")
        else:
            # Create a new sheet with the specified name
            sheet = workbook.create_sheet(title=new_sheet_name)
            break

    workbook.save(file)
    print(f"Sheet '{new_sheet_name}' has been created in the existing file '{file}'.")

def create_sheet_in_new_file(file_name, sheet_name):
    """
    Create a new Excel file with a sheet.
    """
    workbook = openpyxl.Workbook()

    while True:
        new_sheet_name = input("Enter sheet name: ")

        if is_sheet_name_exists(workbook, new_sheet_name):
            print(f"The sheet name '{new_sheet_name}' already exists. Please enter a different name.")
        else:
            # Create a new sheet with the specified name
            sheet = workbook.create_sheet(title=new_sheet_name)
            break

    file = os.path.join(excel_directory, f"{file_name}.xlsx")
    workbook.save(file)
    print(f"Excel file '{file}' with sheet '{new_sheet_name}' has been created.")

def create_excel_file(file_name):
    """
    Create a new Excel file without a sheet.
    """
    file = os.path.join(excel_directory, f"{file_name}.xlsx")

    # Check if the file already exists
    if os.path.exists(file):
        confirmation = input(f"Excel file '{file}' already exists. Do you want to proceed and create a sheet? (y/n): ")
        if confirmation.lower() != 'y':
            print("Operation canceled.")
            return

    # Create a new Excel file
    workbook = openpyxl.Workbook()
    workbook.save(file)
    print(f"Excel file '{file}' has been created.")

def show_files():
    """
    Show existing Excel files in the current directory.
    """
    files = [f for f in os.listdir(excel_directory) if f.endswith('.xlsx')]
    if not files:
        print("No Excel files found in the specified directory.")
    else:
        print("Existing Excel files:")
        for file in files:
            print(file)

# Main program
while True:
    print("\nMenu:")
    print("1. Show Files")
    print("2. Create Sheet")
    print("3. Create Excel File and Sheet")
    print("Enter 'q' to quit")

    choice = input("Enter your choice (1, 2, or 3): ")

    if choice == '1':
        show_files()
    elif choice == '2':
        file_name = input("Enter Excel file name: ")
        file_path = os.path.join(excel_directory, f"{file_name}.xlsx")
        if os.path.exists(file_path):
            create_sheet_in_existing_file(file_path, '')
        else:
            print(f"File '{file_path}' does not exist.")
    elif choice == '3':
        new_file_name = input("Enter new Excel file name: ")
        create_excel_file(new_file_name)
        create_sheet_in_new_file(new_file_name, '')
    elif choice.lower() == 'q':
        break
    else:
        print("Invalid choice. Please choose 1, 2, or 3.")

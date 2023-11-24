from flask import Flask, render_template, request, jsonify
import os
from openpyxl import Workbook, load_workbook
#from openpyxl.formula.parser import parse
#from openpyxl.formula.translate import Translator
#from openpyxl.utils import FORMULAE


app = Flask(__name__)

def calculate_totals(ws, columns):
    # Implement your logic to calculate column_totals and row_totals
    return {}, {}

def evaluate_formula(cell, ws):
    if cell.data_type == 'f':
        # Create a new openpyxl workbook
        wb = Workbook()

        # Create a new worksheet in the new workbook
        ws_copy = wb.active
        ws_copy.title = ws.title

        # Copy all values from the original worksheet to the new worksheet
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            ws_copy.append([cell.value for cell in row])

        # Save the new workbook
        temp_workbook_path = "D:\\PM_MODULE\\temporary_workbook.xlsx"
        wb.save(temp_workbook_path)

        # Reopen the temporary workbook to get the calculated value
        wb = load_workbook(filename=temp_workbook_path, read_only=True)
        ws_copy = wb[ws.title]
        evaluated_formula = ws_copy[cell.coordinate].value
        wb.close()

        return evaluated_formula
    else:
        return cell.value


@app.route('/')
def list_excel_files():
    current_directory = os.path.dirname(os.path.realpath(__file__))
    files = [f for f in os.listdir(current_directory) if f.endswith('.xlsx')]
    return render_template('index.html', files=files)

@app.route('/sheets/<filename>')
def list_sheets(filename):
    current_directory = os.path.dirname(os.path.realpath(__file__))
    file_path = os.path.join(current_directory, filename)
    wb = load_workbook(file_path, read_only=True)
    sheets = wb.sheetnames
    return render_template('sheets.html', filename=filename, sheets=sheets)

@app.route('/columns/<filename>/<sheet_name>', methods=['GET', 'POST'])
def list_columns(filename, sheet_name):
    current_directory = os.path.dirname(os.path.realpath(__file__))
    file_path = os.path.join(current_directory, filename)
    wb = load_workbook(file_path, read_only=True)

    ws = wb[sheet_name]
    columns = [cell.value for cell in ws[1]]

    if request.method == 'POST':
        data_to_store = [request.form[column] for column in columns]
        row_number = ws.max_row + 1

        for col_num, value in enumerate(data_to_store, start=1):
            ws.cell(row=row_number, column=col_num, value=value)

        wb.save(file_path)

    column_totals, row_totals = calculate_totals(ws, columns)
    
    all_data = []
    for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        row_data = []
        for col_num, cell_value in enumerate(row, start=1):
            # If the cell contains a formula, evaluate it
            cell = ws.cell(row=row_number, column=col_num)
            if isinstance(cell_value, str) and cell_value.startswith("="):
                cell_value = evaluate_formula(cell, ws)
            row_data.append(cell_value)
        all_data.append(row_data)

    return render_template('columns.html', filename=filename, sheet_name=sheet_name, columns=columns, column_totals=column_totals, row_totals=row_totals, all_data=all_data)

@app.route('/delete_row/<filename>/<sheet_name>/<int:row_number>', methods=['DELETE'])
def delete_row(filename, sheet_name, row_number):
    current_directory = os.path.dirname(os.path.realpath(__file__))
    file_path = os.path.join(current_directory, filename)

    wb = load_workbook(file_path)
    ws = wb[sheet_name]

    # Check if the row exists before deleting
    if 1 <= row_number <= ws.max_row:
        # Get the values from the deleted row
        deleted_row_values = [cell.value for cell in ws[row_number]]

        # Delete the row in the original sheet
        ws.delete_rows(row_number)

        # Clear the content of the deleted row
        for col in ws.iter_cols(min_col=1, max_col=ws.max_column, min_row=row_number, max_row=row_number):
            for cell in col:
                cell.value = None

        wb.save(file_path)
        wb.close()  # Close the workbook

        # Reopen the workbook to refresh data
        wb = load_workbook(file_path, read_only=True)
        ws = wb[sheet_name]

        # Append the deleted row to the "History" sheet
        history_ws = wb.get_sheet_by_name("History")
        history_ws.append(deleted_row_values)

        wb.save(file_path)
        return jsonify({'message': f'Row {row_number} deleted successfully and moved to History'})
    else:
        return jsonify({'error': f'Row {row_number} does not exist'})

if __name__ == '__main__':
    app.run(debug=True)
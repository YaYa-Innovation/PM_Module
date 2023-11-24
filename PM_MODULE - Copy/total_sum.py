import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime

def update_money_spent(sheet, task_index, completed_amount):
    # Get the current money spent for the task
    current_amount = sheet.cell(row=task_index, column=4).value

    # Deduct the completed amount
    new_amount = current_amount - completed_amount

    # Update the money spent for the task
    sheet.cell(row=task_index, column=4, value=new_amount)

    # Add a record to the history sheet
    history_sheet.append([datetime.now(), f'Task {task_index}', completed_amount])

# Load the Excel workbook
workbook = openpyxl.load_workbook('D:\\PM_MODULE\\your_excel_file.xlsx')

# Select the desired sheet
sheet = workbook['Sheet']  # Replace 'Sheet1' with the actual sheet name

# Add a new sheet for history if it doesn't exist
history_sheet_name = 'History'
if history_sheet_name not in workbook.sheetnames:
    history_sheet = workbook.create_sheet(history_sheet_name)
    history_sheet.append(['Date', 'Task', 'Completed Amount'])
else:
    history_sheet = workbook[history_sheet_name]

# Example: Mark task 3 as completed with a completed amount of 2
completed_task_index = 3
completed_amount = 2
update_money_spent(sheet, completed_task_index, completed_amount)

# Save the changes
workbook.save('D:\\PM_MODULE\\your_excel_file.xlsx')

# Close the workbook
workbook.close()

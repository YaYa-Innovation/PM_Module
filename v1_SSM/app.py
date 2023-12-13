from flask import Flask, render_template, request, redirect, url_for, jsonify, session,flash
import os
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import traceback  # Import traceback module

app = Flask(__name__)
app.secret_key = os.environ.get('FLASK_SECRET_KEY', 'fallback_secret_key')


current_dir = os.path.abspath(os.path.dirname(__file__))

@app.route('/')
def index():
    file_list = get_existing_files()
    return render_template('index.html', file_list=file_list)

@app.route('/create_excel', methods=['POST'])
def create_excel():
    excel_name = request.form['excel_name'].strip()
    if not excel_name:
        return redirect(url_for('index'))

    file_path = os.path.join(current_dir, f'{excel_name}.xlsx')

    if os.path.exists(file_path):
        return redirect(url_for('view_excel', filename=excel_name))

    workbook = Workbook()

    # Create "all_spare" sheet
    all_spare_sheet = workbook.create_sheet(title="all_spare")
    all_spare_sheet.append(["Spare ID", "Spare Name", "Quantity", "Purpose", "Booking"])


    # Create "history" sheet
    history_sheet = workbook.create_sheet(title="history")
    history_sheet.append(["Date", "Event"])


    workbook.save(file_path)

    return redirect(url_for('view_excel', filename=excel_name))

@app.route('/view_excel/<filename>')
def view_excel(filename):
    file_path = os.path.join(current_dir, f'{filename}.xlsx')

    # Check if the file exists
    if not os.path.exists(file_path):
        return "File not found."

    workbook = load_workbook(file_path)
    sheet_names = workbook.sheetnames
    return render_template('view_excel.html', filename=filename, sheet_names=sheet_names)

@app.route('/create_sheet/<filename>', methods=['POST'])
def create_sheet(filename):
    sheet_name = request.form['sheet_name']
    file_path = os.path.join(current_dir, f'{filename}.xlsx')

    # Check if the file exists
    if not os.path.exists(file_path):
        return "File not found."

    workbook = load_workbook(file_path)

    # Check if the sheet already exists
    if sheet_name in workbook.sheetnames:
        return "Sheet already exists."

    # Create a new sheet
    workbook.create_sheet(title=sheet_name)
    sheet = workbook[sheet_name]

    # Define title and headers
    title = "PM Calendar"  # You can modify the title as needed
    headers = ["Spare ID", "Spare Name", "Quantity", "Purpose"]

    # Merge cells for the title in the first row
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    sheet.cell(row=1, column=1, value=title)

    # Write headers to the second row
    sheet.append(headers)

    # Save the workbook
    workbook.save(file_path)

    # Create default sheets
    default_sheets = ['all_spare', 'history']
    for default_sheet in default_sheets:
        if default_sheet not in workbook.sheetnames:
            workbook.create_sheet(title=default_sheet)

    # Save the workbook again after creating default sheets
    workbook.save(file_path)

    return redirect(url_for('view_excel', filename=filename))

@app.route('/create_spare/<filename>/<sheet_name>', methods=['POST'])
def create_spare(filename, sheet_name):
    spare_id = request.form['spare_id']
    spare_name = request.form['spare_name']
    quantity = int(request.form['quantity'])
    purpose = request.form['purpose']

    # Load the existing workbook
    file_path = os.path.join(current_dir, f'{filename}.xlsx')
    wb = load_workbook(file_path)

    # Get the active sheet
    sheet = wb[sheet_name]

    # Check if the Spare ID already exists in the sheet
    for row_num, row in enumerate(sheet.iter_rows(min_row=2, max_col=1, max_row=sheet.max_row), start=2):
        if row and len(row) > 0:  # Check if the row is not empty
            existing_spare_id = row[0].value
            if existing_spare_id == spare_id:
                # Spare ID already exists, update the quantity, booking, and log the operation
                existing_quantity_cell = sheet.cell(row=row_num, column=3)  # Assuming quantity is in the third column
                existing_quantity = existing_quantity_cell.value
                existing_quantity_cell.value = existing_quantity + quantity

                # Update the booking information (modify column index accordingly)
                booking_column_index = 5  # Assuming booking is in the fifth column
                existing_booking_cell = sheet.cell(row=row_num, column=booking_column_index)
                existing_booking_cell.value = "YourBookingInfo"  # Replace with your actual booking information

                # Log the operation in the 'history' sheet
                log_history_operation(wb, filename, f"Updated quantity and booking for Spare ID {spare_id}")

                # Save the workbook
                wb.save(file_path)

                # Redirect to the 'all_spare' sheet
                return redirect(url_for('view_sheet', filename=filename, sheet_name='all_spare'))
        # If Spare ID doesn't exist, create a new entry
    sheet.append([spare_id, spare_name, quantity, purpose,None])

    # Log the operation in the 'history' sheet
    log_history_operation(wb, filename, f"Created new spare with Spare ID {spare_id}")

    # Save the workbook
    wb.save(file_path)

    # Redirect to the 'all_spare' sheet
    return redirect(url_for('view_sheet', filename=filename, sheet_name='all_spare'))

def log_history_operation(wb, operation, details):
    # Get the 'history' sheet or create it if it doesn't exist
    history_sheet = wb['history'] if 'history' in wb.sheetnames else wb.create_sheet('history')

    # Append a new row with the current date, operation, and details
    current_date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    history_sheet.append([current_date, operation, details])

    # Save the changes to the workbook
    current_dir = os.path.dirname(os.path.abspath(__file__))
    wb.save(os.path.join(current_dir, 'Furnace.xlsx'))

@app.route('/view_sheet/<filename>/<sheet_name>')
def view_sheet(filename, sheet_name):
    # Store the current sheet name in the session
    session['current_sheet'] = sheet_name

    file_path = os.path.join(current_dir, f'{filename}.xlsx')

    # Check if the file exists
    if not os.path.exists(file_path):
        return "File not found."

    workbook = load_workbook(file_path, data_only=True)

    # Get the requested sheet
    sheet = workbook[sheet_name]

    # Extract data from the sheet (starting from row 2)
    headers = [cell.value for cell in sheet[2]]
    data = []
    for row in sheet.iter_rows(min_row=3, values_only=True):
        data.append(row)

    return render_template('view_sheet.html', filename=filename, sheet_name=sheet_name, headers=headers, data=data)

def get_existing_files():
    files = [f[:-5] for f in os.listdir(current_dir) if f.endswith('.xlsx')]
    return files

@app.route('/create_task/<filename>/<sheet_name>', methods=['POST'])
def create_task(filename, sheet_name):
    # Access task details from the form using request.form
    task_spare_id = request.form['task_spare_id']
    task_spare_name = request.form['task_spare_name']
    task_qty = request.form['task_qty']
    task_purpose = request.form['task_purpose']

    # Process the task details as needed
    # ...

    # Create a dictionary with the task data
    task_data = {
        'Spare ID': task_spare_id,
        'Spare Name': task_spare_name,
        'Quantity': task_qty,
        'Purpose': task_purpose,
        # Add other fields as needed
    }

    # Load the existing workbook
    file_path = os.path.join(current_dir, f'{filename}.xlsx')
    wb = load_workbook(file_path)

    # Get the active sheet (use the current sheet name stored in the session)
    task_sheet_name = session.get('current_sheet', 'default_sheet_name')
    sheet = wb[task_sheet_name]

    # Find the first empty row in column A
    row_number = sheet.max_row + 1

    # Write the form data to the sheet
    for col_num, (header, value) in enumerate(task_data.items(), start=1):
        sheet.cell(row=row_number, column=col_num, value=value)

    # Save the workbook
    wb.save(file_path)

    # Update the 'all_spare' sheet with the booking quantity
    update_all_spare_sheet(wb, task_spare_id, task_qty)

    # Log the task creation in the history sheet
    log_history_operation(wb, 'Task Created', f'Task for Spare ID {task_spare_id} created with Quantity {task_qty}')

    # Redirect to the view sheet page
    return redirect(url_for('view_sheet', filename=filename, sheet_name=task_sheet_name))


def update_all_spare_sheet(wb, task_spare_id, task_qty):
    all_spare_sheet = wb['all_spare']
    
    # Iterate over rows in the 'all_spare' sheet
    for row in all_spare_sheet.iter_rows(min_row=2, max_col=5, max_row=all_spare_sheet.max_row):
        existing_spare_id, existing_spare_name, existing_qty, purpose, existing_booking_info = [cell.value for cell in row]

        # Check if the spare ID matches the task_spare_id
        if existing_spare_id == task_spare_id:
            try:
                # Update the booking information with the task_qty
                if existing_booking_info is not None:
                    updated_booking_info = str(int(existing_booking_info) + int(task_qty))
                else:
                    # If existing_booking_info is None, set it to the task_qty
                    updated_booking_info = str(task_qty)

                row[4].value = updated_booking_info
            except ValueError:
                # Handle the case where the conversion fails
                print(f"Unable to update booking information for Spare ID {task_spare_id}. Skipping update.")
            break  # Stop searching once the spare ID is found and updated

    # Save the changes to the workbook
    current_dir = os.path.dirname(os.path.abspath(__file__))
    wb.save(os.path.join(current_dir, 'Furnace.xlsx'))


@app.route('/get_spare_name/<filename>/<spare_id>')
def get_spare_name(filename, spare_id):
    file_path = os.path.join(current_dir, f'{filename}.xlsx')

    # Check if the file exists
    if not os.path.exists(file_path):
        return jsonify({"success": False, "spare_name": None})

    workbook = load_workbook(file_path, data_only=True)

    # Get the 'all_spare' sheet
    sheet = workbook['all_spare'] if 'all_spare' in workbook.sheetnames else None

    if sheet:
        # Search for the spare ID and return the corresponding spare name
        for row in sheet.iter_rows(min_row=2, max_col=2, max_row=sheet.max_row):
            existing_spare_id, existing_spare_name = row[0].value, row[1].value
            if existing_spare_id == spare_id:
                return jsonify({"success": True, "spare_name": existing_spare_name})

    return jsonify({"success": False, "spare_name": None})
def update_all_spare_after_completion(wb, task_spare_id, task_qty):
    all_spare_sheet = wb['all_spare']

    # Iterate over rows in the 'all_spare' sheet
    for row in all_spare_sheet.iter_rows(min_row=2, max_col=5, max_row=all_spare_sheet.max_row):
        existing_spare_id, existing_spare_name, existing_qty, purpose, existing_booking_info = [cell.value for cell in row]

        # Check if the spare ID matches the task_spare_id
        if existing_spare_id == task_spare_id:
            try:
                # Subtract the task_qty from the existing_qty
                updated_qty = int(existing_qty) - int(task_qty)

                # Ensure that the quantity does not go below 0
                updated_qty = max(updated_qty, 0)

                # Update the quantity in the 'all_spare' sheet
                row[2].value = updated_qty
            except (ValueError, TypeError) as e:
                # Handle the case where the conversion fails
                print(f"Unable to update quantity for Spare ID {task_spare_id}. Skipping update. Error: {e}")
            break  # Stop searching once the spare ID is found and updated

    # Save the changes to the workbook
    current_dir = os.path.dirname(os.path.abspath(__file__))
    wb.save(os.path.join(current_dir, 'Furnace.xlsx'))

# ... (your other functions)

@app.route('/complete_task/<filename>/<sheet_name>/<int:row_number>', methods=['POST'])
def complete_task(filename, sheet_name, row_number):
    # Load the existing workbook
    current_dir = os.path.dirname(os.path.abspath(__file__))
    wb = load_workbook(os.path.join(current_dir, f'{filename}.xlsx'))

    # Get the active sheet (use the current sheet name stored in the session)
    sheet = wb[sheet_name]

    # Get the spare_id and qty from the specified row
    task_spare_id = sheet.cell(row=row_number, column=1).value
    task_qty = sheet.cell(row=row_number, column=3).value

    # Update 'all_spare'
    update_all_spare_after_completion(wb, task_spare_id, task_qty)

    # Delete the row in the task sheet
    sheet.delete_rows(row_number)

    # Save the changes to the workbook
    wb.save(os.path.join(current_dir, f'{filename}.xlsx'))

    # Redirect to the view sheet page
    return jsonify({"success": True, "message": "Task marked as complete."})


if __name__ == '__main__':
    app.run(debug=True)

from flask import Flask, render_template, request
import os
import openpyxl

app = Flask(__name__)

UPLOAD_FOLDER = os.path.dirname(os.path.abspath(__file__))
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

def create_excel(file_path):
    wb = openpyxl.Workbook()

    # Creating 'All Stock' sheet
    all_stock_sheet = wb.create_sheet("All Stock")
    all_stock_sheet.append(['Spare Name', 'Spare ID', 'Quantity'])

    # Creating 'History' sheet
    history_sheet = wb.create_sheet("History")
    history_sheet.append(['Date', 'Action', 'Spare Name', 'Spare ID', 'Quantity'])

    wb.remove(wb['Sheet'])  # Remove the default sheet

    wb.save(file_path)

@app.route('/')
def index():
    excel_files = get_excel_files()
    return render_template('index.html', excel_files=excel_files)

@app.route('/create_excel', methods=['POST'])
def create_excel_route():
    file_name = request.form['file_name']
    file_path = os.path.join(app.config['UPLOAD_FOLDER'], f'{file_name}.xlsx')
    
    create_excel(file_path)
    
    # Get the relative path for display
    relative_path = os.path.relpath(file_path, app.config['UPLOAD_FOLDER'])
    
    excel_files = get_excel_files()
    return render_template('index.html', excel_files=excel_files, created_file_path=relative_path)

def get_excel_files():
    excel_files = [f for f in os.listdir(app.config['UPLOAD_FOLDER']) if f.endswith('.xlsx')]
    return excel_files

if __name__ == '__main__':
    app.run(debug=True)

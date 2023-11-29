from flask import Flask, render_template, request, redirect, url_for
from openpyxl import load_workbook
from datetime import datetime

app = Flask(__name__)

# Load the main Excel file
tasks_excel_path = "tasks.xlsx"
wb_tasks = load_workbook(tasks_excel_path)
sheet_tasks = wb_tasks.active

# Load the history Excel file
history_excel_path = "history.xlsx"
wb_history = load_workbook(history_excel_path)
sheet_history = wb_history.active

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/all_stock')
def all_stock():
    # Fetch all stock details from the main Excel file
    stock_data = []
    for row in sheet_tasks.iter_rows(min_row=2, values_only=True):
        stock_data.append({
            'spare_name': row[0],
            'spare_id': row[1],
            'qty': row[2],
            'booking': row[3],
        })
    return render_template('all_stock.html', stock_data=stock_data)

@app.route('/add_task', methods=['GET', 'POST'])
def add_task():
    if request.method == 'POST':
        # Handle the form submission and update the main Excel file
        spare_name = request.form['spare']
        spare_id = int(request.form['id'])
        qty_requested = int(request.form['qty'])
        purpose = request.form['purpose']

        # Check if spare exists
        spare_exists = False
        for row in sheet_tasks.iter_rows(min_row=2, max_row=sheet_tasks.max_row, values_only=True):
            if row[0] == spare_name and row[1] == spare_id:
                spare_exists = True
                break

        if spare_exists:
            # Check if qty is available
            available_qty = 0
            for row in sheet_tasks.iter_rows(min_row=2, max_row=sheet_tasks.max_row, values_only=True):
                if row[0] == spare_name and row[1] == spare_id:
                    available_qty = row[2] - row[3]
                    break

            if qty_requested <= available_qty:
                # Update main Excel file and booking qty
                for cell in sheet_tasks['A']:
                    if cell.value == spare_name and cell.offset(column=1).value == spare_id:
                        cell.offset(column=3).value += qty_requested
                        break

                # Record the task in history
                history_row = [datetime.now(), spare_name, spare_id, qty_requested, purpose]
                sheet_history.append(history_row)
                wb_history.save(history_excel_path)

                return redirect(url_for('task_status'))
            else:
                # Handle adjustment if qty is not available
                return render_template('add_task.html', adjustment_msg=f"Adjustment required. Available qty: {available_qty}")
        else:
            return render_template('add_task.html', adjustment_msg="Spare not found.")

    return render_template('add_task.html')
@app.route('/create_spare', methods=['GET', 'POST'])
def create_spare():
    if request.method == 'POST':
        # Handle the form submission and create a new spare
        spare_name = request.form['spare']
        spare_id = int(request.form['id'])

        # Check if spare ID already exists
        for row in sheet_tasks.iter_rows(min_row=2, max_row=sheet_tasks.max_row, values_only=True):
            if row[1] == spare_id:
                return render_template('create_spare.html', error_msg="Spare ID already exists. Please choose a different ID.")

        initial_qty = int(request.form['qty'])  # Initial quantity for the new spare
        purpose = request.form['purpose']

        # Add the new spare to the main Excel file
        sheet_tasks.append([spare_name, spare_id, initial_qty, 0])  # Assuming initial booking is 0
        wb_main.save(main_excel_path)

        # Record the creation of the new spare in history
        history_row = [datetime.now(), spare_name, spare_id, initial_qty, f"New spare created - {purpose}"]
        sheet_history.append(history_row)
        wb_history.save(history_excel_path)

        return redirect(url_for('all_stock'))

    return render_template('create_spare.html')
@app.route('/task_status')
def task_status():
    # Fetch and display task status from the main Excel file
    task_data = []
    for row in sheet_tasks.iter_rows(min_row=2, values_only=True):
        task_data.append({
            'spare_name': row[0],
            'spare_id': row[1],
            'qty': row[2],
            'booking': row[3],
        })
    return render_template('task_status.html', task_data=task_data)

@app.route('/history')
def history():
    # Fetch and display history details from the history Excel file
    history_data = []
    for row in sheet_history.iter_rows(min_row=2, values_only=True):
        history_data.append({
            'timestamp': row[0],
            'spare_name': row[1],
            'spare_id': row[2],
            'qty_requested': row[3],
            'purpose': row[4],
        })
    return render_template('history.html', history_data=history_data)

if __name__ == '__main__':
    app.run(debug=True)

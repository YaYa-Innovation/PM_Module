# D:\STORE_STOCK_MANAGEMENT_SYSTEM\Machine_store_v1\app.py

from flask import Flask, render_template, request, redirect, url_for, jsonify
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import NamedStyle
from datetime import datetime

app = Flask(__name__)

def create_excel_file(file_name):
    file_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), file_name + '.xlsx')
    print(f"Entered create_excel_file function for {file_name}")
    print(f"File Path: {file_path}")

    if os.path.exists(file_path):
        return False, f"File with name '{file_name}' already exists. Please choose a different name."

    try:
        wb = Workbook()

        # Add 'All Stock' sheet with headers
        all_stock_sheet = wb.active
        all_stock_sheet.title = "All Stock"

        headers_all_stock = [
            'Spare Name', 'Spare ID', 'Qty', 'Booking', 'Purpose'
        ]

        for col_num, header in enumerate(headers_all_stock, 1):
            all_stock_sheet.cell(row=1, column=col_num, value=header)

        # No formulas needed, so no need to apply formulas

        # Add 'History' sheet
        history_sheet = wb.create_sheet("History")

        headers_history = [
            'Spare Name', 'Spare ID', 'Qty', 'Booking', 'Purpose', 'Operation', 'Datetime'
        ]

        for col_num, header in enumerate(headers_history, 1):
            history_sheet.cell(row=1, column=col_num, value=header)

        # Apply styles to the 'Datetime' column (assuming it's column 7)
        datetime_column = history_sheet['G']  # Adjust the column letter based on the actual location
        datetime_style = NamedStyle(name='datetime_style', number_format='YYYY-MM-DD HH:MM:SS')
        for cell in datetime_column:
            cell.style = datetime_style

        wb.save(file_path)

        # You can add more sheets here based on user input or a predefined list
        # For example, adding two more sheets named 'Sheet1' and 'Sheet2'
        add_sheet(file_name, 'Sheet1')
        add_sheet(file_name, 'Sheet2')

        print(f"File Created: {file_path}")
        return True, f"Excel file for {file_name} created successfully!"
    except Exception as e:
        print(f"Error Creating File: {e}")
        return False, str(e)

def get_excel_files():
    excel_files = [f for f in os.listdir(os.path.dirname(os.path.abspath(__file__))) if f.endswith('.xlsx')]
    return excel_files

def get_sheet_data(file_name, sheet_name):
    if not file_name.endswith('.xlsx'):
        file_name += '.xlsx'

    file_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), file_name)
    try:
        wb = load_workbook(file_path)
        sheet = wb[sheet_name]
        headers = [cell.value for cell in sheet[1]]
        data = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            data.append(dict(zip(headers, row)))
        return headers, data
    except Exception as e:
        print(f"Error getting sheet data: {e}")
        return [], []

@app.route('/')
def index():
    print("Entered index route")
    excel_files = get_excel_files()
    return render_template('index.html', message='', excel_files=excel_files)

@app.route('/add_area', methods=['POST'])
def add_area():
    print("Entered add_area route")
    area_name = request.form['area_name']
    success, error_message = create_excel_file(area_name)
    return jsonify(success=success, message=error_message)

@app.route('/open_excel/<file_name>')
def open_excel(file_name):
    file_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), file_name)
    try:
        wb = load_workbook(file_path)
        sheets = wb.sheetnames
        return render_template('sheets_viewer.html', file_name=file_name, sheets=sheets)
    except Exception as e:
        return f"Error opening Excel file: {e}"

@app.route('/open_sheet/<file_name>/<sheet_name>')
def open_sheet(file_name, sheet_name):
    headers, data = get_sheet_data(file_name, sheet_name)
    return render_template('sheet_viewer.html', file_name=file_name, sheet_name=sheet_name, headers=headers, data=data)

@app.route('/add_data/<file_name>/<sheet_name>', methods=['GET', 'POST'])
def add_data(file_name, sheet_name):
    if request.method == 'POST':
        form_data = {
            'spare_name': request.form.get('spare_name'),
            'spare_id': request.form.get('spare_id'),
            'qty': request.form.get('qty'),
            'booking': request.form.get('booking'),
            'purpose': request.form.get('purpose'),
        }

        success, message = add_task(file_name, sheet_name, form_data)
        if success:
            return redirect(url_for('view_sheet_data', file_name=file_name, sheet_name=sheet_name))
        else:
            return render_template('add_data_form.html', file_name=file_name, sheet_name=sheet_name, message=message)

    return render_template('add_data_form.html', file_name=file_name, sheet_name=sheet_name, message='')

# ...

def add_task(file_name, sheet_name, form_data):
    try:
        file_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), file_name)
        wb = load_workbook(file_path)
        sheet = wb[sheet_name]

        row_num = 2
        while sheet.cell(row=row_num, column=1).value is not None:
            row_num += 1

        sheet.cell(row=row_num, column=1, value=form_data['spare_name'])
        sheet.cell(row=row_num, column=2, value=form_data['spare_id'])
        sheet.cell(row=row_num, column=3, value=0)  # Set the quantity to 0 in "Qty" column
        sheet.cell(row=row_num, column=4, value=form_data['qty'])  # Set the booking quantity in "Booking" column
        sheet.cell(row=row_num, column=5, value=form_data['booking'])
        sheet.cell(row=row_num, column=6, value=form_data['purpose'])

        wb.save(file_path)

        update_all_stock(file_name, form_data['spare_name'], form_data['spare_id'], form_data['qty'], form_data['booking'])
        update_history(file_name, form_data['spare_name'], form_data['spare_id'], form_data['qty'], form_data['booking'], form_data['purpose'], 'In')

        return True, 'Task added successfully.'
    except Exception as e:
        return False, str(e)

# ...



def add_sheet(file_name, sheet_name):
    file_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), file_name)
    try:
        wb = load_workbook(file_path)
        if sheet_name not in wb.sheetnames:
            new_sheet = wb.create_sheet(sheet_name)

            headers = [
                'Spare Name', 'Spare ID', 'Qty', 'Booking', 'Purpose'
            ]

            for col_num, header in enumerate(headers, 1):
                new_sheet.cell(row=1, column=col_num, value=header)

            wb.save(file_path)
            return True, f"Sheet '{sheet_name}' added successfully with headers."
        else:
            return False, f"Sheet '{sheet_name}' already exists. Please choose a different name."
    except Exception as e:
        return False, str(e)

@app.route('/add_sheet_route/<file_name>', methods=['POST'])
def add_sheet_route(file_name):
    sheet_name = request.form['sheet_name']
    success, message = add_sheet(file_name, sheet_name)
    return jsonify(success=success, message=message)

@app.route('/view_sheet_data/<file_name>/<sheet_name>')
def view_sheet_data(file_name, sheet_name):
    headers, data = get_sheet_data(file_name, sheet_name)
    return render_template('sheet_data_viewer.html', file_name=file_name, sheet_name=sheet_name, headers=headers, data=data)

def update_all_stock(file_name, spare_name, spare_id, qty, booking):
    try:
        file_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), file_name)
        wb = load_workbook(file_path)
        sheet = wb['All Stock']

        for row in range(2, sheet.max_row + 1):
            if sheet.cell(row=row, column=1).value == spare_name and sheet.cell(row=row, column=2).value == spare_id:
                current_qty = sheet.cell(row=row, column=3).value
                current_booking = sheet.cell(row=row, column=4).value
                new_booking = current_booking + int(booking)  # Update the "Booking" column
                sheet.cell(row=row, column=4, value=new_booking)

                wb.save(file_path)
                return True, 'All Stock updated successfully.'

        return False, 'Item not found in All Stock.'
    except Exception as e:
        return False, str(e)




def update_history(file_name, spare_name, spare_id, qty, booking, purpose, operation):
    try:
        file_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), file_name)
        wb = load_workbook(file_path)
        history_sheet = wb['History']

        history_row_num = 2
        while history_sheet.cell(row=history_row_num, column=1).value is not None:
            history_row_num += 1

        history_sheet.cell(row=history_row_num, column=1, value=spare_name)
        history_sheet.cell(row=history_row_num, column=2, value=spare_id)
        history_sheet.cell(row=history_row_num, column=3, value=qty)
        history_sheet.cell(row=history_row_num, column=4, value=booking)
        history_sheet.cell(row=history_row_num, column=5, value=purpose)
        history_sheet.cell(row=history_row_num, column=6, value=operation)
        history_sheet.cell(row=history_row_num, column=7, value=datetime.now())

        wb.save(file_path)
    except Exception as e:
        print(f"Error updating history: {e}")

@app.route('/add_spare/<file_name>', methods=['GET', 'POST'])
def add_spare(file_name):
    if request.method == 'POST':
        spare_name = request.form.get('spare_name')
        spare_id = request.form.get('spare_id')
        qty = request.form.get('qty')
        booking = 0
        purpose = request.form.get('purpose')

        success, message = update_all_stock(file_name, spare_name, spare_id, qty, booking)
        if success:
            update_history(file_name, spare_name, spare_id, qty, booking, purpose, 'In')
            return redirect(url_for('view_sheet_data', file_name=file_name, sheet_name='All Stock'))
        else:
            return render_template('add_spare_form.html', file_name=file_name, message=message)

    return render_template('add_spare_form.html', file_name=file_name, message='')

def move_to_history(file_name, sheet_name, row_num, operation):
    try:
        file_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), file_name)
        wb = load_workbook(file_path)
        sheet = wb[sheet_name]
        history_sheet = wb['History']

        task_data = [cell.value for cell in sheet[row_num]]

        history_row_num = 2
        while history_sheet.cell(row=history_row_num, column=1).value is not None:
            history_row_num += 1

        history_sheet.cell(row=history_row_num, column=1, value=task_data[0])
        history_sheet.cell(row=history_row_num, column=2, value=task_data[1])
        history_sheet.cell(row=history_row_num, column=3, value=task_data[2])
        history_sheet.cell(row=history_row_num, column=4, value=task_data[3])
        history_sheet.cell(row=history_row_num, column=5, value=task_data[4])
        history_sheet.cell(row=history_row_num, column=6, value=operation)
        history_sheet.cell(row=history_row_num, column=7, value=datetime.now())

        sheet.delete_rows(row_num)

        wb.save(file_path)
        return True, 'Task moved to History successfully.'
    except Exception as e:
        return False, str(e)

@app.route('/complete_task/<file_name>/<sheet_name>/<int:row_num>', methods=['POST'])
def complete_task(file_name, sheet_name, row_num):
    try:
        spare_name = request.json['spare_name']
        spare_id = request.json['spare_id']
        qty = request.json['qty']
        booking = request.json['booking']
        purpose = request.json['purpose']

        success, message = move_to_history(file_name, sheet_name, row_num, 'Complete')
        if success:
            update_history(file_name, spare_name, spare_id, qty, booking, purpose, 'Complete')
            return jsonify(success=True, message=message)
        else:
            return jsonify(success=False, message=message)
    except Exception as e:
        return jsonify(success=False, message=str(e))

def get_spare_name_from_all_stock(file_name, spare_id):
    file_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), file_name)
    
    try:
        wb = load_workbook(file_path)
        all_stock_sheet = wb['All Stock']

        # Iterate through rows to find spare ID
        for row in all_stock_sheet.iter_rows(min_row=2, values_only=True):
            if row[1] == spare_id:  # Assuming the spare ID is in the second column
                spare_name = row[0]  # Assuming the spare name is in the first column
                wb.close()
                return spare_name

        wb.close()
        return None
    except Exception as e:
        print(f"Error getting spare name from All Stock: {e}")
        return None

# Modify the existing /get_spare_name route to use the new function
@app.route('/get_spare_name/<file_name>/<spare_id>', methods=['GET'])
def get_spare_name(file_name, spare_id):
    spare_name = get_spare_name_from_all_stock(file_name, spare_id)
    if spare_name:
        return jsonify({'success': True, 'spare_name': spare_name})
    else:
        return jsonify({'success': False, 'error': 'Spare ID not found'}), 404


if __name__ == '__main__':
    app.run(debug=True)

import openpyxl
from datetime import datetime
import os

# Get the script's directory
script_dir = os.path.dirname(os.path.realpath(__file__))

# Construct the Excel file path relative to the script's directory
excel_file_path = os.path.join(script_dir, 'your_excel_file.xlsx')

# Check if the file exists
if not os.path.exists(excel_file_path):
    # If the file doesn't exist, create a new workbook
    workbook = openpyxl.Workbook()

    # Access the default sheet and add sample data
    sheet = workbook.active
    sheet['A1'] = 'Date'
    sheet['B1'] = 'Due Date'
    sheet['A2'] = '2023-12-06'
    sheet['B2'] = 'Due today'
    sheet['A3'] = '2023-12-07'
    sheet['B3'] = 'Due today1'
    sheet['A4'] = '2023-12-08'
    sheet['B4'] = 'Due today2'
    sheet['A5'] = '2023-12-09'
    sheet['B5'] = 'Due today3'

    # Save the workbook with the specified file name
    workbook.save(excel_file_path)
    workbook.close()

# Open the Excel file
workbook = openpyxl.load_workbook(excel_file_path)

# Specify the sheet name
sheet_name = 'Sheet'

# Specify the columns
date_column = 'A'
due_date_column = 'B'

# Access the sheet
sheet = workbook[sheet_name]

# Find the row with the maximum date
max_row = max(sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True), key=lambda row: row[0])

# Print the values in the two columns for the row with the maximum date
max_date = max_row[0]
due_date = max_row[1]
print(f"Max Date: {max_date}, Due Date: {due_date}")

# Close the Excel file
workbook.close()

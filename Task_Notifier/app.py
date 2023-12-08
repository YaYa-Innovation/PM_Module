import os
from flask import Flask, render_template, request, redirect, url_for
from openpyxl import Workbook, load_workbook

app = Flask(__name__)

current_dir = os.path.abspath(os.path.dirname(__file__))

@app.route('/')
def index():
    file_list = get_existing_files()
    return render_template('index.html', file_list=file_list)

@app.route('/create_excel', methods=['POST'])
def create_excel():
    excel_name = request.form['excel_name'].strip()  # Remove leading/trailing whitespaces
    if not excel_name:
        # Handle the case where the provided name is empty
        return redirect(url_for('index'))

    file_path = os.path.join(current_dir, f'{excel_name}.xlsx')

    # Check if the file already exists
    if os.path.exists(file_path):
        return redirect(url_for('view_excel', filename=excel_name))

    # If not, create a new Excel file
    workbook = Workbook()
    workbook.save(file_path)

    return redirect(url_for('view_excel', filename=excel_name))

@app.route('/view_excel/<filename>')
def view_excel(filename):
    file_path = os.path.join(current_dir, f'{filename}.xlsx')

    # Check if the file exists
    if not os.path.exists(file_path):
        return "File not found."

    workbook = load_workbook(file_path)
    sheet_names = workbook.sheetnames
    return render_template('view_excel.html', filename=filename, sheet_names=sheet_names)

@app.route('/create_sheet/<filename>', methods=['POST'])
def create_sheet(filename):
    sheet_name = request.form['sheet_name']
    file_path = os.path.join(current_dir, f'{filename}.xlsx')

    # Check if the file exists
    if not os.path.exists(file_path):
        return "File not found."

    workbook = load_workbook(file_path)

    # Check if the sheet already exists
    if sheet_name in workbook.sheetnames:
        return "Sheet already exists."

    # Create a new sheet
    workbook.create_sheet(title=sheet_name)
    sheet = workbook[sheet_name]

    # Define title and headers
    title = "PM Calendar"  # You can modify the title as needed
    headers = ["Machine No", "Machine Name", "Machine Part", "Job number", "spare id", "spare name",
               "qty", "frequency in days", "time required in hrs", "man power requirement",
               "cost", "total cost", "last changed date", "due date"]

    # Merge cells for the title in the first row
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    sheet.cell(row=1, column=1, value=title)

    # Write headers to the second row
    sheet.append(headers)

    # Save the workbook
    workbook.save(file_path)

    return redirect(url_for('view_excel', filename=filename))

@app.route('/view_sheet/<filename>/<sheet_name>')
def view_sheet(filename, sheet_name):
    file_path = os.path.join(current_dir, f'{filename}.xlsx')

    # Check if the file exists
    if not os.path.exists(file_path):
        return "File not found."

    workbook = load_workbook(file_path, data_only=True)

    # Get the requested sheet
    sheet = workbook[sheet_name]

    # Extract data from the sheet (starting from row 2)
    headers = [cell.value for cell in sheet[2]]
    data = []
    for row in sheet.iter_rows(min_row=3, values_only=True):
        data.append(row)

    return render_template('view_sheet.html', filename=filename, sheet_name=sheet_name, headers=headers, data=data)




@app.route('/complete_row/<filename>/<sheet_name>/<int:row_index>')
def complete_row(filename, sheet_name, row_index):
    file_path = os.path.join(current_dir, f'{filename}.xlsx')

    # Check if the file exists
    if not os.path.exists(file_path):
        return "File not found."

    workbook = load_workbook(file_path)

    # Get the requested sheet
    sheet = workbook[sheet_name]

    # Perform actions for completing the row (you can customize this part)
    # For now, let's mark the row as completed by updating a specific cell
    completed_cell = sheet.cell(row=row_index + 2, column=len(sheet[2]) + 1)
    completed_cell.value = "Completed"

    # Save the workbook
    workbook.save(file_path)

    return redirect(url_for('view_sheet', filename=filename, sheet_name=sheet_name))

def get_existing_files():
    files = [f[:-5] for f in os.listdir(current_dir) if f.endswith('.xlsx')]
    return files

@app.route('/create_task/<filename>/<sheet_name>', methods=['POST'])
def create_task(filename, sheet_name):
    # Access task details from the form using request.form
    task_machine_no = request.form['task_machine_no']
    task_machine_name = request.form['task_machine_name']
    task_machine_part = request.form['task_machine_part']
    task_job_number = request.form['task_job_number']
    task_spare_id = request.form['task_spare_id']
    task_spare_name = request.form['task_spare_name']
    task_qty = request.form['task_qty']
    task_frequency = request.form['task_frequency']
    task_time_required = request.form['task_time_required']
    task_manpower = request.form['task_manpower']
    task_cost = request.form['task_cost']
    task_total_cost = request.form['task_total_cost']
    task_last_changed_date = request.form['task_last_changed_date']
    task_due_date = request.form['task_due_date']

    # Process the task details as needed
    # ...

    # Create a dictionary with the task data
    task_data = {
        'Machine No': task_machine_no,
        'Machine Name': task_machine_name,
        'Machine Part': task_machine_part,
        'Job Number': task_job_number,
        'Spare ID': task_spare_id,
        'Spare Name': task_spare_name,
        'Quantity': task_qty,
        'Frequency in Days': task_frequency,
        'Time Required in Hours': task_time_required,
        'Manpower Requirement': task_manpower,
        'Cost': task_cost,
        'Total Cost': task_total_cost,
        'Last Changed Date': task_last_changed_date,
        'Due Date': task_due_date,
    }

    # Load the existing workbook
    file_path = os.path.join(current_dir, f'{filename}.xlsx')
    wb = load_workbook(file_path)

    # Get the active sheet
    sheet = wb[sheet_name]

    # Find the first empty row in column A
    row_number = sheet.max_row + 1

    # Write the form data to the sheet
    for col_num, (header, value) in enumerate(task_data.items(), start=1):
        sheet.cell(row=row_number, column=col_num, value=value)

    # Save the workbook
    wb.save(file_path)

    # Redirect to the view sheet page
    return redirect(url_for('view_sheet', filename=filename, sheet_name=sheet_name))


if __name__ == '__main__':
    app.run(debug=True)

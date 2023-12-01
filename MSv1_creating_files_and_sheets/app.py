# D:\STORE_STOCK_MANAGEMENT_SYSTEM\Machine_store_v1\app.py

from flask import Flask, render_template, request, redirect, url_for, jsonify
import os
from openpyxl import Workbook, load_workbook

app = Flask(__name__)

def create_excel_file(file_name):
    file_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), file_name + '.xlsx')
    print(f"Entered create_excel_file function for {file_name}")
    print(f"File Path: {file_path}")

    if os.path.exists(file_path):
        return False, f"File with name '{file_name}' already exists. Please choose a different name."

    try:
        wb = Workbook()
        default_sheet = wb.active  # Get the default first sheet
        default_sheet.title = "All Stock"  # Rename the default sheet
        wb.save(file_path)

        # You can add more sheets here based on user input or a predefined list
        # For example, adding two more sheets named 'Sheet1' and 'Sheet2'
        add_sheet(file_name, 'Sheet1')
        add_sheet(file_name, 'Sheet2')

        print(f"File Created: {file_path}")
        return True, f"Excel file for {file_name} created successfully!"
    except Exception as e:
        print(f"Error Creating File: {e}")
        return False, str(e)

def get_excel_files():
    excel_files = [f for f in os.listdir(os.path.dirname(os.path.abspath(__file__))) if f.endswith('.xlsx')]
    return excel_files

@app.route('/')
def index():
    print("Entered index route")
    excel_files = get_excel_files()
    return render_template('index.html', message='', excel_files=excel_files)

@app.route('/add_area', methods=['POST'])
def add_area():
    print("Entered add_area route")
    area_name = request.form['area_name']
    success, error_message = create_excel_file(area_name)
    return jsonify(success=success, message=error_message)

@app.route('/open_excel/<file_name>')
def open_excel(file_name):
    file_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), file_name)
    try:
        wb = load_workbook(file_path)
        sheets = wb.sheetnames
        return render_template('sheets_viewer.html', file_name=file_name, sheets=sheets)
    except Exception as e:
        return f"Error opening Excel file: {e}"

def add_sheet(file_name, sheet_name):
    file_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), file_name)
    try:
        wb = load_workbook(file_path)
        if sheet_name not in wb.sheetnames:
            wb.create_sheet(sheet_name)
            wb.save(file_path)
            return True, f"Sheet '{sheet_name}' added successfully."
        else:
            return False, f"Sheet '{sheet_name}' already exists. Please choose a different name."
    except Exception as e:
        return False, str(e)

@app.route('/add_sheet/<file_name>/<sheet_name>', methods=['POST'])
def add_sheet_route(file_name, sheet_name):
    success, message = add_sheet(file_name, sheet_name)
    return jsonify(success=success, message=message)

if __name__ == '__main__':
    app.run(debug=True)

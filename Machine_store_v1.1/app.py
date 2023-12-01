# D:\STORE_STOCK_MANAGEMENT_SYSTEM\Machine_store_v1\app.py

from flask import Flask, render_template, request, redirect, url_for, jsonify
import os
from openpyxl import Workbook, load_workbook
from datetime import datetime  # Import the datetime module

app = Flask(__name__)


def create_excel_file(file_name):
    file_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), file_name + '.xlsx')
    print(f"Entered create_excel_file function for {file_name}")
    print(f"File Path: {file_path}")

    if os.path.exists(file_path):
        return False, f"File with name '{file_name}' already exists. Please choose a different name."

    try:
        wb = Workbook()
        
        # Add 'All Stock' sheet with headers
        all_stock_sheet = wb.active
        all_stock_sheet.title = "All Stock"

        headers = [
            'Machine Number', 'Machine Name', 'Machine Part', 'Category', 'Spare ID', 'Spare Name',
            'Qty', 'Frequency in Days', 'Time Required in Hrs', 'Man Power Requirement', 
            'Man Hour', 'Cost', 'Total Cost'
        ]

        for col_num, header in enumerate(headers, 1):
            all_stock_sheet.cell(row=1, column=col_num, value=header)

        # Apply formulas for column 11 and 13 within a specific range
        for row_num in range(2, 10):  # Adjust the range based on your actual requirements
            all_stock_sheet.cell(row=row_num, column=11, value=f'=SUM(J{row_num}*I{row_num})')
            all_stock_sheet.cell(row=row_num, column=13, value=f'=SUM(G{row_num}*K{row_num})')

        # Add another sheet named 'History'
        history_sheet = wb.create_sheet("History")

        wb.save(file_path)

        # You can add more sheets here based on user input or a predefined list
        # For example, adding two more sheets named 'Sheet1' and 'Sheet2'
        add_sheet(file_name, 'Sheet1')
        add_sheet(file_name, 'Sheet2')

        print(f"File Created: {file_path}")
        return True, f"Excel file for {file_name} created successfully!"
    except Exception as e:
        print(f"Error Creating File: {e}")
        return False, str(e)

def get_excel_files():
    excel_files = [f for f in os.listdir(os.path.dirname(os.path.abspath(__file__))) if f.endswith('.xlsx')]
    return excel_files

def get_sheet_data(file_name, sheet_name):
    # Check if the file name already contains the extension
    if not file_name.endswith('.xlsx'):
        file_name += '.xlsx'

    file_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), file_name)
    try:
        wb = load_workbook(file_path)
        sheet = wb[sheet_name]
        headers = [cell.value for cell in sheet[1]]
        data = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            data.append(dict(zip(headers, row)))
        #print(f"Headers: {headers}")
        #print(f"Data: {data}")
        return headers, data
    except Exception as e:
        print(f"Error getting sheet data: {e}")
        return [], []



@app.route('/')
def index():
    print("Entered index route")
    excel_files = get_excel_files()
    return render_template('index.html', message='', excel_files=excel_files)

@app.route('/add_area', methods=['POST'])
def add_area():
    print("Entered add_area route")
    area_name = request.form['area_name']
    success, error_message = create_excel_file(area_name)
    return jsonify(success=success, message=error_message)

@app.route('/open_excel/<file_name>')
def open_excel(file_name):
    file_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), file_name)
    try:
        wb = load_workbook(file_path)
        sheets = wb.sheetnames
        return render_template('sheets_viewer.html', file_name=file_name, sheets=sheets)
    except Exception as e:
        return f"Error opening Excel file: {e}"

@app.route('/open_sheet/<file_name>/<sheet_name>')
def open_sheet(file_name, sheet_name):
    headers, data = get_sheet_data(file_name, sheet_name)
    return render_template('sheet_viewer.html', file_name=file_name, sheet_name=sheet_name, headers=headers, data=data)

@app.route('/add_data/<file_name>/<sheet_name>', methods=['GET', 'POST'])
def add_data(file_name, sheet_name):
    if request.method == 'POST':
        form_data = {
            'name': request.form.get('name'),
            'age': request.form.get('age'),
        }

        success, message = add_task(file_name, sheet_name, form_data)
        if success:
            # Redirect to the sheet_data_viewer route to refresh the data
            return redirect(url_for('view_sheet_data', file_name=file_name, sheet_name=sheet_name))
        else:
            return render_template('add_data_form.html', file_name=file_name, sheet_name=sheet_name, message=message)

    return render_template('add_data_form.html', file_name=file_name, sheet_name=sheet_name, message='')

def add_task(file_name, sheet_name, form_data):
    try:
        file_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), file_name)
        wb = load_workbook(file_path)
        sheet = wb[sheet_name]

        # Find the first empty row in the sheet
        row_num = 2
        while sheet.cell(row=row_num, column=1).value is not None:
            row_num += 1

        # Add the task data to the sheet
        sheet.cell(row=row_num, column=1, value=form_data['name'])
        sheet.cell(row=row_num, column=2, value=form_data['age'])

        # Save the changes to the workbook
        wb.save(file_path)

        return True, 'Task added successfully.'
    except Exception as e:
        return False, str(e)


def add_sheet(file_name, sheet_name):
    file_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), file_name)
    try:
        wb = load_workbook(file_path)
        if sheet_name not in wb.sheetnames:
            wb.create_sheet(sheet_name)
            wb.save(file_path)
            return True, f"Sheet '{sheet_name}' added successfully."
        else:
            return False, f"Sheet '{sheet_name}' already exists. Please choose a different name."
    except Exception as e:
        return False, str(e)

@app.route('/add_sheet/<file_name>/<sheet_name>', methods=['POST'])
def add_sheet_route(file_name, sheet_name):
    success, message = add_sheet(file_name, sheet_name)
    return jsonify(success=success, message=message)


@app.route('/view_sheet_data/<file_name>/<sheet_name>')
def view_sheet_data(file_name, sheet_name):
    headers, data = get_sheet_data(file_name, sheet_name)
    return render_template('sheet_data_viewer.html', file_name=file_name, sheet_name=sheet_name, headers=headers, data=data)

if __name__ == '__main__':
    app.run(debug=True)
